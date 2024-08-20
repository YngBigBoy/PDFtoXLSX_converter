import os
import sys
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QPushButton, QFileDialog, QMessageBox
import camelot
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill
import gc


class PDFtoExcelConverter(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        self.layout = QVBoxLayout()

        self.btn_single = QPushButton('Конвертировать PDF в отдельные файлы Excel', self)
        self.btn_single.clicked.connect(self.convert_single)
        self.layout.addWidget(self.btn_single)

        self.btn_multiple = QPushButton('Конвертировать PDF в один файл Excel', self)
        self.btn_multiple.clicked.connect(self.convert_multiple)
        self.layout.addWidget(self.btn_multiple)

        self.setLayout(self.layout)
        self.setWindowTitle('Конвертер PDF в Excel')
        self.show()

    def convert_single(self):
        options = QFileDialog.Options()
        files, _ = QFileDialog.getOpenFileNames(self, "Выберите PDF файлы для конвертации", "", "PDF Files (*.pdf)", options=options)
        exept = False
        if files:
            destination_folder = QFileDialog.getExistingDirectory(self, "Выберите папку назначения")
            if destination_folder:
                for pdf_file in files:
                    try:
                        self.convert_pdf_to_excel(pdf_file, destination_folder)
                    except Exception as e:
                        exept = True
                        QMessageBox.warning(self, "Ошибка", f"Не удалось конвертировать {os.path.basename(pdf_file)}: {str(e)}")
                    finally:
                        gc.collect()
                if not exept:
                    QMessageBox.information(self, "Успех", f"Конвертация PDF-файла(ов) завершена успешно!")

    def convert_multiple(self):
        options = QFileDialog.Options()
        files, _ = QFileDialog.getOpenFileNames(self, "Выберите PDF файлы для конвертации", "", "PDF Files (*.pdf)", options=options)
        if files:
            save_path, _ = QFileDialog.getSaveFileName(self, "Сохранить файл Excel", "", "Excel Files (*.xlsx)", options=options)
            if save_path:
                try:
                    self.convert_pdfs_to_one_excel(files, save_path)
                except Exception as e:
                    QMessageBox.warning(self, "Ошибка", f"Не удалось конвертировать PDF файлы: {str(e)}")
                finally:
                    QMessageBox.information(self, "Успех", f"Конвертация PDF в {os.path.basename(save_path)} завершена успешно!")
                    gc.collect()

    def convert_pdf_to_excel(self, pdf_path, destination_folder):
        # Извлечение таблиц с помощью Camelot
        tables = camelot.read_pdf(pdf_path, pages='all')

        excel_path = os.path.join(destination_folder, os.path.splitext(os.path.basename(pdf_path))[0] + '.xlsx')
        wb = Workbook()
        ws = wb.active
        ws.title = 'Лист1'

        start_row = 1
        previous_header = None
        first_table = True

        if not tables:
            ws.cell(row=start_row, column=1).value = f"{os.path.basename(pdf_path)} не содержит таблиц или поврежден"
            wb.save(excel_path)
            return

        for table in tables:
            df = table.df
            if df.empty or df.dropna(how='all').empty:
                continue

            # Убираем возможный сдвиг заголовков вниз
            df.columns = df.iloc[0]  # Присваиваем первую строку в качестве заголовка
            df = df.drop(0).reset_index(drop=True)  # Убираем первую строку, которая теперь стала заголовком

            current_header = list(df.columns)

            # Если заголовок изменился, добавляем отступ между таблицами
            if current_header != previous_header:
                if not first_table:
                    start_row += 3  # отступаем на 3 строки между разными таблицами из одного файла
                first_table = False

            # Записываем данные таблицы
            for r in dataframe_to_rows(df, index=False, header=(start_row == 1 or (previous_header != current_header))):
                for c_idx, value in enumerate(r, 1):
                    cell = ws.cell(row=start_row, column=c_idx, value=value)

                    # Применяем форматирование
                    if start_row == 1 or (r == list(df.columns)):
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='center')

                start_row += 1

            previous_header = current_header

        # Автоматическая настройка ширины столбцов
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Получаем букву столбца
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = adjusted_width

        wb.save(excel_path)

    def convert_pdfs_to_one_excel(self, pdf_files, excel_path):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Все таблицы'

        start_row = 1

        for pdf_file in pdf_files:
            tables = camelot.read_pdf(pdf_file, pages='all')
            pdf_filename = os.path.basename(pdf_file)
            current_header = None
            first_table = True

            # Добавляем одну пустую строку перед каждым файлом
            if start_row != 1:
                start_row += 1

            if not tables:
                ws.cell(row=start_row, column=1).value = f"{pdf_filename} не содержит таблиц или поврежден"
                start_row += 2
                continue

            for table in tables:
                df = table.df
                if df.empty or df.dropna(how='all').empty:
                    continue

                df.columns = df.iloc[0]  # Присваиваем первую строку в качестве заголовка
                df = df.drop(0).reset_index(drop=True)  # Убираем первую строку, которая теперь стала заголовком

                df.insert(0, 'Файл', pdf_filename)

                # Проверяем заголовок текущей таблицы
                if list(df.columns) != current_header:
                    if current_header != list(df.columns):
                        start_row += 3  # отступаем на 3 строки между разными таблицами из одного файла

                    if first_table:
                        start_row += 2
                        first_table = False

                # Записываем данные таблицы
                for r in dataframe_to_rows(df, index=False, header=(start_row == 1 or (current_header != list(df.columns)))):
                    for c_idx, value in enumerate(r, 1):
                        cell = ws.cell(row=start_row, column=c_idx, value=value)

                        if start_row == 1 or (r == list(df.columns)):
                            cell.font = Font(bold=True)
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                        else:
                            cell.alignment = Alignment(horizontal='left', vertical='center')
                    start_row += 1
                current_header = list(df.columns)

                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter  # Получаем букву столбца
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = max_length + 2
                    ws.column_dimensions[column].width = adjusted_width
        wb.save(excel_path)


if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = PDFtoExcelConverter()
    sys.exit(app.exec_())